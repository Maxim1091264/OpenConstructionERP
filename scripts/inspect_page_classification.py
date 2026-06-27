from __future__ import annotations

import sys
from pathlib import Path
from openpyxl import load_workbook


def main(argv):
    if len(argv) < 2:
        print("Usage: python scripts/inspect_page_classification.py <xlsx-path>")
        return 2
    p = Path(argv[1])
    if not p.exists():
        print(f"File not found: {p}")
        return 2
    wb = load_workbook(p, read_only=True)
    ws = wb['Pages']
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    total = len(rows)
    count_relevant = sum(1 for r in rows if (r[3] and str(r[3]).lower()=='yes'))
    print(f"Total pages: {total}, Estimate-relevant pages: {count_relevant}")
    return 0


if __name__ == '__main__':
    raise SystemExit(main(sys.argv))
