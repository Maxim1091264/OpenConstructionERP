"""Adapter for XLSX-based quantity and pricing tables."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable
from uuid import uuid4

from openpyxl import load_workbook

from app.modules.ai_smeta_ru.domain.models import ExtractedObject


class XlsxAdapter:
    """Reads workbook sheets and produces structured row objects."""

    def extract(self, file_path: str, artifact_id: str | None = None) -> list[ExtractedObject]:
        workbook = load_workbook(filename=file_path, data_only=True)
        artifact_id = artifact_id or str(uuid4())
        objects: list[ExtractedObject] = []

        for sheet in workbook.worksheets:
            for row_index, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                values = [str(cell).strip() for cell in row if cell is not None]
                if not values:
                    continue
                raw_text = " | ".join(values)
                object_id = f"xlsx-{sheet.title}-{row_index}-{uuid4().hex}"
                objects.append(
                    ExtractedObject(
                        object_id=object_id,
                        artifact_id=artifact_id,
                        object_type="table_row",
                        raw_text=raw_text,
                        normalized_text=raw_text,
                        confidence=0.95,
                    )
                )

        return objects
