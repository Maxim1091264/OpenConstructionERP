from __future__ import annotations

from typing import Iterable
from openpyxl import Workbook

from app.modules.ai_smeta_ru.services.document_intelligence_ru_service import PageClassification


class PageClassificationExcelExport:
    def export(self, pages: Iterable[PageClassification], output_path: str) -> str:
        wb = Workbook()
        pages_sheet = wb.active
        pages_sheet.title = "Pages"
        pages_sheet.append([
            "Page Number",
            "Page Type",
            "Discipline Guess",
            "Is Estimate Relevant",
            "Confidence",
            "Reason",
            "Text Preview",
        ])

        relevant_rows = []
        for p in pages:
            pages_sheet.append([
                p.page_number,
                p.page_type,
                p.discipline_guess or "",
                "yes" if p.is_estimate_relevant else "no",
                p.confidence,
                p.reason,
                p.extracted_text_preview,
            ])
            if p.is_estimate_relevant:
                relevant_rows.append(p)

        rel_sheet = wb.create_sheet(title="EstimateRelevantPages")
        rel_sheet.append([
            "Page Number",
            "Page Type",
            "Discipline Guess",
            "Confidence",
            "Reason",
            "Text Preview",
        ])
        for p in relevant_rows:
            rel_sheet.append([
                p.page_number,
                p.page_type,
                p.discipline_guess or "",
                p.confidence,
                p.reason,
                p.extracted_text_preview,
            ])

        log_sheet = wb.create_sheet(title="ProcessingLog")
        log_sheet.append(["Stage", "Status", "Notes"])
        log_sheet.append(["PageClassification", "completed", f"Processed {len(list(pages))} pages"])

        wb.save(output_path)
        return output_path
