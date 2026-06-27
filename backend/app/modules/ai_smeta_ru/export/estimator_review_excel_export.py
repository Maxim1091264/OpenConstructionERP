"""Exports estimator review data to Excel for the AI-Smeta-RU prototype."""

from __future__ import annotations

from typing import Iterable

from openpyxl import Workbook

from app.modules.ai_smeta_ru.domain.models import QuantityItem, WorkItem


class EstimatorReviewExcelExport:
    """Exports work items and quantity items into a review-friendly workbook."""

    def export(
        self,
        work_items: Iterable[WorkItem],
        quantity_items: Iterable[QuantityItem],
        output_path: str,
    ) -> str:
        workbook = Workbook()
        work_items_sheet = workbook.active
        work_items_sheet.title = "WorkItems"
        work_items_sheet.append([
            "WorkItem ID",
            "Description",
            "Unit",
            "Status",
            "Estimated Category",
            "Source Reference",
        ])

        for work_item in work_items:
            work_items_sheet.append([
                work_item.work_item_id,
                work_item.description,
                work_item.unit or "",
                work_item.status,
                work_item.estimated_category or "",
                work_item.source_reference or "",
            ])

        quantity_sheet = workbook.create_sheet(title="QuantityItems")
        quantity_sheet.append([
            "QuantityItem ID",
            "WorkItem ID",
            "Quantity",
            "Unit",
            "Measurement Basis",
            "Source Reference",
        ])

        for quantity_item in quantity_items:
            quantity_sheet.append([
                quantity_item.quantity_id,
                quantity_item.work_item_id or "",
                quantity_item.quantity if quantity_item.quantity is not None else "",
                quantity_item.unit or "",
                quantity_item.measurement_basis or "",
                quantity_item.source_reference or "",
            ])

        log_sheet = workbook.create_sheet(title="ProcessingLog")
        log_sheet.append(["Stage", "Status", "Notes"])
        log_sheet.append(["Export", "completed", "Estimate review package generated"])

        workbook.save(output_path)
        return output_path
